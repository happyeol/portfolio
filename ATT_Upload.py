#ATT_Upload.py
import ATT_Util

from openpyxl import load_workbook

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys

#패키지 업로드를 위한 엑셀파일 불러오기
wb = load_workbook('ATT_Package Upload.xlsx')

#변수 값 초기화
ws_info = wb['Info']
model = ws_info['C2'].value
version = ws_info['C3'].value
path = ws_info['C4'].value
Event_num = ws_info['C5'].value
ATT_ID = ws_info['C6'].value
ATT_PW = ws_info['C7'].value
EP_ID = ws_info['C8'].value
EP_PW = ws_info['C9'].value

ws_file = wb['Filename']

Fileinfo_list = []
for i in range(2,ws_file.max_row+1):
    Fileinfo_list.append([ws_file['B'+str(i)].value, ws_file['C'+str(i)].value, ws_file['D'+str(i)].value])
#print(Fileinfo_list)

#서버 접속
url = "https://xdme.wireless.att.com/jsp/login/login.jsp"

driver = webdriver.Chrome('chromedriver')
#print(type(driver))
driver.implicitly_wait(300)
driver.get(url)

#서버 로그인
ATT_Util.login_input(driver,"LOGIN",ATT_ID)
ATT_Util.login_input(driver,"PASSWORD",ATT_PW)
driver.find_element(By.NAME,'subBotton').send_keys(Keys.ENTER)

#업로드 필요한 패키지 확인
Upload_list = ATT_Util.Chcek_Package(driver,model,version,Fileinfo_list)

#패키지 업로드 수행
if Upload_list:
    while True:
        Failed_list = ATT_Util.Upload_Package(driver,model,version,Upload_list)
        if Failed_list: #업로드 실패한 경우
            Upload_list = Failed_list
            continue
        else: #업로드 성공한 경우
            print("모든 패키지가 업로드 되었습니다.")
            break
else:
    print("모든 패키지가 업로드된 상태입니다.")